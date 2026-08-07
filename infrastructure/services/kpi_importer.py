import openpyxl

from infrastructure.models import KPI


def import_kpis_from_excel(file):

    workbook = openpyxl.load_workbook(
        file,
        data_only=True
    )

   # Remove previous KPI configuration
    # KPI.objects.all().delete()

    imported = 0

    # ---------------------------------------------------------
    # Read every sheet
    # Sheet name = KPI category
    # ---------------------------------------------------------

    for sheet_name in workbook.sheetnames:

        category = sheet_name.strip().upper()

        worksheet = workbook[sheet_name]

        rows = worksheet.iter_rows(
            values_only=True
        )

        # -----------------------------------------------------
        # Skip header row
        # -----------------------------------------------------

        next(rows, None)

        # -----------------------------------------------------
        # Execution order starts from 1
        # for each sheet/category
        # -----------------------------------------------------

        execution_order = 1

        for row in rows:

            if not row:
                continue

            # -------------------------------------------------
            # KPI Name
            # -------------------------------------------------

            kpi_name = row[0]

            if not kpi_name:
                continue

            # -------------------------------------------------
            # Linux Command
            # -------------------------------------------------

            linux_command = (
                row[1]
                if len(row) > 1
                else None
            )

            # -------------------------------------------------
            # Windows Command
            # -------------------------------------------------

            windows_command = (
                row[2]
                if len(row) > 2
                else None
            )

            # -------------------------------------------------
            # Save KPI
            # -------------------------------------------------

            KPI.objects.update_or_create(

                name=str(
                    kpi_name
                ).strip(),

                category=category,

                defaults={

                    "linux_command": (
                        str(
                            linux_command
                        ).strip()
                        if linux_command
                        else None
                    ),

                    "windows_command": (
                        str(
                            windows_command
                        ).strip()
                        if windows_command
                        else None
                    ),

                    "execution_order": execution_order,

                    "enabled": True,

                }
            )

            execution_order += 1

            imported += 1

    return imported